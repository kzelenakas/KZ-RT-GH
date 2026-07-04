"""Preprocess loader: dedupes identical sheets, applies clean+split, no PII in fixtures."""

import openpyxl
import pytest

from app.revision_mining.preprocess import extract_atomic_items, load_unique_sheets


@pytest.fixture
def synthetic_xlsx(tmp_path):
    path = tmp_path / "synthetic_rrr.xlsx"
    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "SheetA"
    ws1.append(["REVISION REQUEST"])
    ws1.append(["1. Correct the subject ZIP code.\n2. Provide support for the site adjustment."])
    ws1.append(["N/A"])

    ws2 = wb.create_sheet("SheetB-Duplicate")
    ws2.append(["REVISION REQUEST"])
    ws2.append(["1. Correct the subject ZIP code.\n2. Provide support for the site adjustment."])
    ws2.append(["N/A"])

    ws3 = wb.create_sheet("SheetC-Unique")
    ws3.append(["REVISION REQUEST"])
    ws3.append(["Add the missing comparable photo."])

    wb.save(path)
    return path


def test_load_unique_sheets_dedupes_identical_content(synthetic_xlsx):
    sheets = load_unique_sheets(synthetic_xlsx)
    assert len(sheets) == 2  # SheetA and SheetB-Duplicate are identical; SheetC-Unique differs


def test_extract_atomic_items_cleans_splits_and_dedupes(synthetic_xlsx):
    items = extract_atomic_items(synthetic_xlsx)
    assert len(items) == 3  # 2 items from the one kept duplicate sheet + 1 from the unique sheet
    assert any("ZIP code" in i for i in items)
    assert any("site adjustment" in i for i in items)
    assert any("comparable photo" in i for i in items)
    assert not any(i.strip() == "N/A" for i in items)
