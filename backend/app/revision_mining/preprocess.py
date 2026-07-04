from __future__ import annotations

import hashlib
from pathlib import Path

import openpyxl

from .clean_split import clean_revision_text, split_revision_text


def _sheet_content_hash(ws) -> str:
    """Hash a sheet's non-blank first-column values (skipping the header row).
    Used to detect byte-identical duplicate sheets (seen in the real export:
    two 'Master' tabs were exact copies of each other)."""
    h = hashlib.sha256()
    for row in ws.iter_rows(min_row=2, values_only=True):
        value = (row[0] or "").strip() if row and row[0] else ""
        if value:
            h.update(value.encode("utf-8", "ignore"))
    return h.hexdigest()


def load_unique_sheets(xlsx_path: Path) -> list:
    """Load all worksheets, dropping any whose content is byte-identical to
    an earlier sheet (keeps the first occurrence)."""
    wb = openpyxl.load_workbook(xlsx_path, read_only=True)
    seen_hashes: set[str] = set()
    kept = []
    for ws in wb.worksheets:
        digest = _sheet_content_hash(ws)
        if digest in seen_hashes:
            continue
        seen_hashes.add(digest)
        kept.append(ws)
    return kept


def extract_atomic_items(xlsx_path: Path) -> list[str]:
    """Load unique sheets, clean each REVISION REQUEST cell, split bundled
    cells into atomic action items. Returns a flat list of cleaned strings."""
    items: list[str] = []
    for ws in load_unique_sheets(xlsx_path):
        for row in ws.iter_rows(min_row=2, values_only=True):
            raw = row[0] if row else None
            cleaned = clean_revision_text(raw or "")
            if cleaned is None:
                continue
            items.extend(split_revision_text(cleaned))
    return items
